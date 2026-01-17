import re
from openpyxl import load_workbook

# =============================
# FUNCION Clean_id
# Elimina caracteres no numericos de un documento
# cc76.888.560
75888560
# =============================

# De aqui en adelante el codigo es la Logica del Negocio

def clean_id(value):
    # Eliminar caractertes no numericos de un documento
    if value is None:
        return ""
    return re.sub(r'\D','',str(value))

# =======================
#  🤖Funcion merge_name🤖
#    Une nombre y apellido en un solo campo 
# =======================

def merge_name(name, lastname):
    if name is None:
        name = ""
    if lastname is None:
        lastname = ""
    return f"{name} {lastname}".strip()

def process_excel(path):
    # Acceso a la hoja llamada "Datos"
    wb = load_workbook(path)
    ws = wb["Datos"]
    # Recorre todas las filas desde la fila 2
    for row in range(2, ws.max_row+1):
    # Columna D: identificador limpio
        ws[f"D{row}"] = clean_id(ws[f"A{row}"].value)
        # Columna E: Nombre Completo
        ws[f"E{row}"] = merge_name(
        ws[f"B{row}"].value,
        ws[f"C{row}"].value
            )
        # Guarda los cambios en el mismo archivo
        wb.save(path)

# Hasta aqui va la Logica del Negocio

# De aqui en adelante todo el codigo hace parte de un archivoo Controlador

def process_excel_safe(path):
    try:
        process_excel(path)
        return True, "Archivo procesado correctamente"
    except PermissionError:
        return(
            False, "El archivo Excel esta abierto.\n"
            "Por favor, cierralo e intenta nuevamente."
        )
    except KeyError:
        return False, "Hoja 'Datos' no encontrada"
    except Exception as e:
        return False, f"Error inesperado: {str(e)}"