# processor.py
# Lógica de negocio: operaciones sobre Excel

import re
from openpyxl import load_workbook


def clean_id(value):
    # Eliminar caractertes no numericos de un documento
    if value is None:
        return ""
    return re.sub(r'\D','',str(value))

def merge_name(name, lastname):
    if name is None:
        name = ""
    if lastname is None:
        lastname = ""
    return f"{name} {lastname}".strip()

def ejecutar_accion(instruccion,pash):
    # Abre el archivo de ejemplo
    
    wb = load_workbook(pash)
    ws = wb.active

    # Funcion para limpiar
    if instruccion["action"] == "clean_id":
        col = instruccion["column"]
        for fila in range(2, ws.max_row + 1):
            ws[f"{col}{fila}"] = ''.join(filter(str.isdigit, str(ws[f"{col}{fila}"].value)))

    # Funcion para unir
    elif instruccion["action"] == "merge_name":
        for fila in range(2, ws.max_row + 1):
            nombre = ws["A" + str(fila)].value or ""
            apellido = ws["B" + str(fila)].value or ""
            ws["C" + str(fila)] = f"{nombre} {apellido}".strip()

    wb.save(pash)

    